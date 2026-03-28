import os
import pathlib
import datetime

import pandas
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import xlwings as xw

from Setup.Mail import AconexMailType
from Setup.config import config

class MailColumn:
    def __init__(self, column_name, datatype, mandatory : bool = False):
        self.__heading : str = column_name
        self.__datatype : str = datatype
        self.__isMandatory : bool = mandatory or "*" == self.__heading[-1] #asterisk marks mandatory
        self.__value : str | list[str] | datetime.datetime | bool = None

    def heading_name(self) -> str:
        return self.__heading

    def is_mandatory(self) -> bool:
        return self.__isMandatory

    def get_length(self) -> int:
        if self.__datatype == "SINGLE_LINE_TEXT":
            return 25
        elif self.__datatype == "DATE":
            return 13
        elif self.__datatype == "name":
            return 20
        elif self.__datatype == "namelist" or self.__datatype == "filelist":
            return 30
        elif self.__datatype == "confidential":
            return 8
        elif self.__datatype == "html":
            return 80
        else:
            return 30

    def set_column_formatting(self, workbook, worksheet, colindex : int, engine : str):
        col_len = self.get_length()

        if engine == "xlsxwriter":
            cellformat = workbook.add_format({
                'text_wrap': True
            })
            if self.is_mandatory():
                cellformat.set_bold()

            if self.__datatype == "DATE":
                cellformat =  workbook.add_format({'number_format': 'dd/mm/yyyy'})

            elif self.__datatype == "confidential":
                worksheet.data_validation(
                    3,colindex,worksheet.max_row,3,
                    {
                        'validate': 'list',
                        'source': ['TRUE', 'FALSE']
                    }
                )

            worksheet.set_column(colindex, colindex, col_len, cellformat)

        elif engine == "openpyxl":
            col_letter = get_column_letter((colindex + 1))
            worksheet.column_dimensions[col_letter].width = col_len
            cellref = col_letter + "2"
            worksheet[cellref].alignment = Alignment(wrap_text=True)

            if self.is_mandatory():
                worksheet[cellref].font = Font(bold=True)
            else:
                worksheet[cellref].font = Font(bold=False)

            if self.__datatype == "DATE":
                pass #this is not doing anything
                # for row in range(3, worksheet.max_row):
                #     cell = worksheet.cell(row=row, column=colindex)
                #     cell.number_format = "dd/mm/yyyy"
                #
                # dvalidation = DataValidation(type="date", allow_blank=True)
                # worksheet.add_data_validation(dvalidation)
                # colrange = '{letter}3:{letter}1000'.format(
                #     letter=col_letter)  # apply to whole col except first two rows
                # dvalidation.add(colrange)

            elif self.__datatype == "confidential":
                dvalidation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
                worksheet.add_data_validation(dvalidation)
                colrange = '{letter}3:{letter}1000'.format(letter=col_letter) #apply to whole col except first two rows
                dvalidation.add(colrange)

        else:
            config.logger.warning("No formatting configured for engine %s" % engine)

    #given the inputted cell value, check this is valid for this column
    def validate(self, rowval : str) -> bool:
        if self.__isMandatory and pandas.isna(rowval):
            config.logger.error("Mandatory column '{0}' is empty".format(self.heading_name()))
            return False

        if self.__datatype == "SINGLE_LINE_TEXT" or self.__datatype == "name":
            self.__value = "" if pandas.isna(rowval) else rowval
            config.logger.debug(self.__value)
            return True

        elif self.__datatype == "namelist" or self.__datatype == "filelist":
            self.__value = rowval.split("; ") if pandas.notna(rowval) else []
            config.logger.debug(self.__value)
            return True

        elif self.__datatype == "DATE":
            if pandas.isna(rowval):
                self.__value = None
            else:
                try:
                    rowdatetime = rowval.date()
                    self.__value = rowdatetime
                except ValueError:
                    config.logger.error("Invalid date format {0}'".format(rowval))
                    return False

            config.logger.debug(self.__value)
            return True
        elif self.__datatype == "html":
            #TODO - how to validate this?
            self.__value = rowval
            return True

        elif self.__datatype == "confidential":
            self.__value = (rowval == True)

            config.logger.debug(self.__value)
            return True

#Universal mail cols
umailCols : list[MailColumn] = [
    MailColumn("Mail Number (Leave blank for auto-number)", "SINGLE_LINE_TEXT"),
    MailColumn("Reply to Mail Number (Leave blank for new thread)", "SINGLE_LINE_TEXT"),
    MailColumn("To Names (Separate by semi-colon)*", "namelist"),
    MailColumn("From Name (Leave blank for NM)", "name"),
    MailColumn("CC Names (Separate by semi-colon)", "namelist"),
    MailColumn("Subject*", "SINGLE_LINE_TEXT"),
    MailColumn("Response Required Date", "DATE"),
    MailColumn("Mail Body", "html"),
    MailColumn("Attachment file names (Separate by semi-colon)", "filelist"),
    MailColumn("Confidential", "confidential")
]

def main(templateupdate : bool = True):
    global FOLDERPATH
    FOLDERPATH = str(pathlib.Path(__file__).parent.resolve())

    if templateupdate:
        createTemplate() #create excel template for project's mail setup

    mailDFDict = importExcel() #get data from excel template

    if mailDFDict:
        registerToAconex(mailDFDict) #register mail data onto aconex


def createTemplate():
    getProjectMailData() #get the mail form field setup for this project
    config.logger.info("Project Mail Template created")

def getProjectMailData():
    filename: str = FOLDERPATH + "\\" + config.project().projectCodePrefix() + "Mail_Template.xlsx"
    if not os.path.exists(filename):
        addInstructions(filename)

    mailTypes : list[AconexMailType] = config.mailtypes()
    assert mailTypes is not None

    for mtype in mailTypes:
        writer = pandas.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists="replace")
        sheetname : str = mtype.corrtypeid() #have to use ID as some mail type names are too long

        pfs = mtype.projectfields()
        #TODO - we need to get any attributes that apply to all mail as well
        morecols = [MailColumn(h, dt, m) for (h, dt, m) in pfs]
        #config.logger.debug("{} : {}".format(mtype.typename(), mtype.projectfields()))

        mailCols = umailCols + morecols
        config.info("Creating sheet for %s" % sheetname)
        #Write Mail Type as title of the sheet
        titledf = pandas.DataFrame(data={"Mail Type": [mtype.typename()], "ID": mtype.corrtypeid()})
        titledf.to_excel(writer,
                           sheet_name=sheetname,
                           header=False,
                           startrow=0,
                           index=False)
        writer.close()
        writer = pandas.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists="overlay")
        #Create data frame of header column names
        headersdf = pandas.DataFrame(columns=[mc.heading_name() for mc in mailCols])
        #write headings to excel sheet
        headersdf.to_excel(writer,
                           sheet_name=sheetname,
                           header=True,
                           startrow=1,
                           index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheetname]

        for index, column in enumerate(headersdf):
            mcol : MailColumn = mailCols[index]
            mcol.set_column_formatting(workbook, worksheet, index, writer.engine)

        writer.close()

    config.logger.info("Adding External General Use label using xlwings")
    wb = xw.Book(filename)
    labelinfo = wb.api.SensitivityLabel.CreateLabelInfo()
    labelinfo.AssignmentMethod = 2
    labelinfo.Justification = 'init'
    labelinfo.LabelId = '428e5717-1df4-456d-8576-f827ff2d20e1'
    wb.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    wb.close()

def addInstructions(filename : str):
    writer = pandas.ExcelWriter(filename, mode='w')
    workbook = writer.book
    worksheet = workbook.add_worksheet('Instructions')
    #TODO - add instructions I CBA
    writer.close()

def importExcel() -> dict | None:
    filename: str = FOLDERPATH + "\\" + config.project().projectCodePrefix() + "Mail_Template.xlsx"
    try:
        excelDataDFDict = pandas.read_excel(open(filename, 'rb'), sheet_name=None, skiprows=[0]) #by specifying no sheet name, it pulls all sheets
    except FileNotFoundError:
        config.logger.error("Could not find template at %s" % filename)
        return None

    #Sanitise - remove empty sheets with no data
    excelDataDFDict = dict(filter(lambda kv : kv[1].size > 0, excelDataDFDict.items()))
    return excelDataDFDict

def registerToAconex(mailDFDict : dict):

    for key, maildf in mailDFDict.items():
        maildf = maildf.reindex(sorted(maildf.columns), axis=1) #sort cols alphabetically
        mailcorrid : str = key
        mtypes : list[AconexMailType] = list(filter(lambda mt: mt.corrtypeid() == mailcorrid , config.mailtypes()))
        assert len(mtypes) == 1 #Validate the mail corr id exists on project
        mtype = mtypes[0]

        pfs = mtype.projectfields()
        morecols = [MailColumn(h, dt, m) for (h, dt, m) in pfs]
        mailCols = umailCols + morecols
        mailCols.sort(key=lambda mcol: mcol.heading_name()) #sort alphabetically

        #check sheet headings match cols
        headings = list(maildf.columns.values)
        colheadernames = [col.heading_name() for col in mailCols]

        assert colheadernames == headings

        #for each row in the sheet
        for i, row in maildf.iterrows():

            for j, mcol in enumerate(mailCols):
                rowval = row.iloc[j]
                config.logger.debug("{col}: {val}".format(col=mcol.heading_name(), val=rowval))
                if mcol.validate(rowval):
                    pass
                else:
                    continue
